from openpyxl import load_workbook

import config


def getUserByNumber(number="", chat_id=""):
    workbook = load_workbook(filename=config.DB_PATH)
    sheet_ranges = workbook['Полная БД']
    column_a = sheet_ranges['A']
    column_b = sheet_ranges['B']
    column_c = sheet_ranges['C']
    column_d = sheet_ranges['D']
    column_e = sheet_ranges['E']
    column_f = sheet_ranges['F']
    column_g = sheet_ranges['G']
    column_h = sheet_ranges['H']
    column_i = sheet_ranges['I']
    column_m = sheet_ranges['M']

    data = []

    if number != "":
        for i in range(len(column_i)):
            if column_i[i].value == number:
                column_m[i].value = str(chat_id)  # update chat id
                workbook.save(config.DB_PATH)
                data = [column_a[i].value,column_b[i].value,column_c[i].value,column_d[i].value,column_e[i].value,column_f[i].value,column_g[i].value,column_h[i].value,column_i[i].value, i+1]
        return data
    else:
        for i in range(len(column_i)):
            if column_m[i].value == str(chat_id):
                data = [column_a[i].value, column_b[i].value, column_c[i].value, column_d[i].value, column_e[i].value,
                        column_f[i].value, column_g[i].value, column_h[i].value, column_i[i].value]

        return data


def deleteRow(row: int):
    workbook = load_workbook(filename=config.DB_PATH)

    ws = workbook.active

    ws.delete_rows(row)

    workbook.save(config.DB_PATH)


def createRow() -> int:
    workbook = load_workbook(filename=config.DB_PATH)

    ws = workbook.active

    first_column = len([cell.value for cell in ws["A"] if cell.value is not None])
    ws.insert_rows(first_column + 1)

    workbook.save(config.DB_PATH)
    return first_column + 1


def addPartRow(row: int, key: str, value: str):
    workbook = load_workbook(filename=config.DB_PATH)

    ws = workbook.active

    if key == "FIO":
        ws.cell(row=row, column=1, value=value)
    elif key == "group":
        ws.cell(row=row, column=2, value=value)
    elif key == "gavrilov":
        ws.cell(row=row, column=3, value=value)
    elif key == "fmba":
        ws.cell(row=row, column=4, value=value)
    elif key == "sum":
        ws.cell(row=row, column=5, value=value)
    elif key == "last_gavrilov":
        ws.cell(row=row, column=6, value=value)
    elif key == "last_fmba":
        ws.cell(row=row, column=7, value=value)
    elif key == "social":
        ws.cell(row=row, column=8, value=value)
    elif key == "phone":
        ws.cell(row=row, column=9, value=value)
    elif key == "chat_id":
        ws.cell(row=row, column=13, value=value)

    workbook.save(config.DB_PATH)


def addRow(data: list) -> int:
    workbook = load_workbook(filename=config.DB_PATH)

    ws = workbook.active

    first_column = len([cell.value for cell in ws["A"] if cell.value is not None])
    ws.insert_rows(first_column+1)
    for i in range(len(data)):
        ws.cell(row=first_column+1, column=i+1, value=str(data[i]))

    workbook.save(config.DB_PATH)
    return first_column+1


def getRow(row: int) -> list:
    workbook = load_workbook(filename=config.DB_PATH)

    ws = workbook.active

    return [cell.value for cell in ws[row]]


def addData(row_index,name,role,gavrilova,fmba,summ,gavrilova_date,fmba_date,contact,phone,chatId,readyToCome,came):
    workbook = load_workbook(filename=config.DB_PATH)
    sheet = workbook.active
    sheet.insert_rows(row_index)

    sheet[f'A{row_index}'] = name
    sheet[f'B{row_index}'] = role
    sheet[f'C{row_index}'] = gavrilova
    sheet[f'D{row_index}'] = fmba
    sheet[f'E{row_index}'] = summ
    sheet[f'F{row_index}'] = gavrilova_date
    sheet[f'G{row_index}'] = fmba_date
    sheet[f'H{row_index}'] = contact
    sheet[f'I{row_index}'] = phone
    sheet[f'j{row_index}'] = chatId
    sheet[f'l{row_index}'] = readyToCome
    sheet[f'k{row_index}'] = came


    workbook.save(config.DB_PATH)
